from openpyxl import load_workbook,Workbook
from datetime import datetime
class Shop:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        self.input_field = dict()
        self.data = dict()
        self.sold_data = dict()
        
    def read_excel(self):
        wb = load_workbook(self.file_path)
        ws = wb.active
        for i in range(1,ws.max_column+1):
            self.input_field[ws.cell(row=1,column=i).value] = ws.cell(row=2,column=i).value    
        for i in range(1, ws.max_column + 1):  # +1 qo'shildi
            self.data[ws.cell(row=1, column=i).value] = []
            for j in range(3, ws.max_row + 1):  # +1 qo'shildi
                self.data[ws.cell(row=1, column=i).value].append(ws.cell(row=j, column=i).value)   
        
    def add_products(self):
        while True:
            print("Mahsulotlar xususiyatlari ushbu tartibda kiritlsin...") ,print((self.input_field))
            print("Chiqish uchun 'exit' so'zini kiriting")
            nomi = input(" Qanday Mahsulot qo'shmoqchisiz: ")
            if nomi == '0':
                wb.save(self.file_path)
                break
            if nomi == 'exit':
                self.main()
            miqdor = input("Mahsulot miqdorini kiriting: ")
            narx = input("Mahsulot narxini kiriting: ")
            rang = input("Mahsulot rangini kiriting:(IXTIYORIY) ")
            date = datetime.now().strftime("%Y-%m-%d")
            type_ = input("Mahsulot turini kiriting: ")  
            self.data["name"].append(nomi)
            self.data["quenty"].append(miqdor)
            self.data["price"].append(narx)
            self.data["date"].append(date)
            self.data["color"].append(rang)
            self.data["type"].append(type_)  
            with open('file.txt','a+') as f:
                f.write(f' Qoshildi =>>>{nomi} \n Miqdori =>>>{miqdor}ta(kg)\n Narxi =>>> {narx}\n Qoshilgan vaqti =>>>{date}\n Turi =>>> {type_}\n_______________________________\n')
            print("Mahsulot muvaffaqiyatli qo'shildi.")
            wb = Workbook()
            ws = wb.active

            titles = ["name", "quenty", "price", "date", "color", "type"]  
            for i, title in enumerate(titles, start=1):
                ws.cell(row=1, column=i, value=title)
            for i in range(len(self.data["name"])):
                for j, title in enumerate(titles, start=1):
                    ws.cell(row=i+2, column=j, value=self.data[title][i])

    def sell_products(self):
        print("Chiqish uchun exit so'zini kiriting ")
        while True:
            mahsulot = (self.data['name'])
            print(f"Bizda ushbu mahsulotlar bor >>>{mahsulot}")
            nomi = input("Qanday mahsulot olasiz? ")
            if nomi in self.data["name"]:
                idx = self.data["name"].index(nomi)
                print(f"{nomi} mahsuloti mavjud. Miqdori: {self.data['quenty'][idx]} ta")
                miqdor = int(input("Mahsulot miqdorini kiriting: "))
                self.data['quenty'][idx] = int(self.data['quenty'][idx])
                self.data['quenty'][idx] -= int(miqdor)
                narxii = int(self.data['price'][idx])
                jami = narxii*miqdor
            else:
                print(f"{nomi} mahsuloti mavjud emas." )   
            if nomi == 'exit':
                self.main()
                date = datetime.now().strftime("%Y-%m-%d")
            with open('sold.txt','a+')  as file:
                file.write(f" Sotildi>>>{nomi}\n Miqdori>>>{miqdor}\n Narxi>>{narxii} JamiSumma>> {jami}_________________________\n")   
            print("Mahsulot muvaffaqiyatli sotildi.")
    def hisobot(self):
        select = input('1.Sotilgan mahsulotlar\n2.Qoshilgan mahsulotlar')
        if select == '1':
            with open('sold.txt','r') as file:
                print(file.read())
        if select == '2':
            with open('file.txt', 'r') as f:
                print(f.read())    
    def main(self):
        while True:
            sora = input('1.Mahsulot qoshish\n2.Mahsulot sotish\n3.Xisobot')
            if sora == '1':
                self.add_products()
            elif sora == '2':
                self.sell_products() 
            elif sora == '3':
                self.hisobot()
            elif sora == '0':  
                wb.save('shooping.xlsx')
                break    
            else:
                print("To'g'ri malumot kiriting? ")                            
shop = Shop('shooping.xlsx')
shop.read_excel()
shop.main()


